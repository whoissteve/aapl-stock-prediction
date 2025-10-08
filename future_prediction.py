# Future Predictions & Excel Export
# Predict AAPL closing prices for next 100 days and export to Excel

import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime, timedelta
from sklearn.preprocessing import MinMaxScaler
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import warnings
warnings.filterwarnings('ignore')

class FuturePricePredictor:
    """
    Make future predictions and export to professional Excel report
    """
    
    def __init__(self, symbol="AAPL"):
        self.symbol = symbol
        self.data = None
        self.models = {}
        self.feature_scaler = MinMaxScaler()
        self.feature_columns = []
        
    def prepare_data(self, period="5y"):
        """Step 1: Download and prepare historical data"""
        print("=" * 70)
        print("STEP 1: PREPARING HISTORICAL DATA")
        print("=" * 70)
        
        print(f"\n📥 Downloading {period} of {self.symbol} data...")
        stock = yf.Ticker(self.symbol)
        data = stock.history(period=period)
        data.reset_index(inplace=True)
        
        print(f"✅ Downloaded {len(data)} days of historical data")
        print(f"📅 Date range: {data['Date'].iloc[0].date()} to {data['Date'].iloc[-1].date()}")
        
        # Add technical indicators (features)
        print("\n Creating technical indicators...")
        data['MA_5'] = data['Close'].rolling(5).mean()
        data['MA_10'] = data['Close'].rolling(10).mean()
        data['MA_20'] = data['Close'].rolling(20).mean()
        data['Price_Change'] = data['Close'].pct_change()
        data['Volume_MA'] = data['Volume'].rolling(10).mean()
        data['Volume_Ratio'] = data['Volume'] / data['Volume_MA']
        data['High_Low_Pct'] = (data['High'] - data['Low']) / data['Close']
        data['Close_lag_1'] = data['Close'].shift(1)
        data['Close_lag_2'] = data['Close'].shift(2)
        
        # Remove NaN values
        data.dropna(inplace=True)
        
        self.data = data
        print(f"✅ Prepared {len(data)} rows with {data.shape[1]} features")
        
        return data
    
    def train_models(self):
        """Step 2: Train prediction models"""
        print("\n" + "=" * 70)
        print("STEP 2: TRAINING PREDICTION MODELS")
        print("=" * 70)
        
        # Define feature columns
        self.feature_columns = ['Open', 'High', 'Low', 'Volume', 'MA_5', 'MA_10', 
                               'MA_20', 'Price_Change', 'Volume_MA', 'Volume_Ratio',
                               'High_Low_Pct', 'Close_lag_1', 'Close_lag_2']
        
        X = self.data[self.feature_columns]
        y = self.data['Close']
        
        # Split data (80% train, 20% test)
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42, shuffle=False
        )
        
        print(f"\n Training set: {len(X_train)} days")
        print(f" Test set: {len(X_test)} days")
        
        # Scale features
        X_train_scaled = self.feature_scaler.fit_transform(X_train)
        X_test_scaled = self.feature_scaler.transform(X_test)
        
        # Train Linear Regression
        print("\n🤖 Training Linear Regression model...")
        lr_model = LinearRegression()
        lr_model.fit(X_train_scaled, y_train)
        self.models['Linear_Regression'] = lr_model
        
        # Evaluate Linear Regression
        lr_pred = lr_model.predict(X_test_scaled)
        lr_r2 = r2_score(y_test, lr_pred)
        lr_mae = mean_absolute_error(y_test, lr_pred)
        
        print(f"    R² Score: {lr_r2:.4f} ({lr_r2*100:.2f}% accuracy)")
        print(f"    MAE: ${lr_mae:.2f}")
        
        # Train Random Forest
        print("\n Training Random Forest model...")
        rf_model = RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1)
        rf_model.fit(X_train_scaled, y_train)
        self.models['Random_Forest'] = rf_model
        
        # Evaluate Random Forest
        rf_pred = rf_model.predict(X_test_scaled)
        rf_r2 = r2_score(y_test, rf_pred)
        rf_mae = mean_absolute_error(y_test, rf_pred)
        
        print(f"    R² Score: {rf_r2:.4f} ({rf_r2*100:.2f}% accuracy)")
        print(f"    MAE: ${rf_mae:.2f}")
        
        print("\n✅ Both models trained successfully!")
        
        return {
            'Linear_Regression': {'r2': lr_r2, 'mae': lr_mae},
            'Random_Forest': {'r2': rf_r2, 'mae': rf_mae}
        }
    
    def predict_future(self, days_ahead=100):
        """Step 3: Predict future prices for next N days"""
        print("\n" + "=" * 70)
        print(f"STEP 3: PREDICTING NEXT {days_ahead} DAYS")
        print("=" * 70)
        
        # Get last known data point
        last_data = self.data.iloc[-1:].copy()
        last_date = last_data['Date'].values[0]
        last_close = last_data['Close'].values[0]
        
        print(f"\n Starting from: {pd.to_datetime(last_date).date()}")
        print(f" Last known price: ${last_close:.2f}")
        
        # Containers for predictions
        predictions = {
            'Linear_Regression': [],
            'Random_Forest': [],
            'dates': []
        }
        
        current_date = pd.to_datetime(last_date)
        
        print(f"\n🔮 Generating {days_ahead} future predictions...")
        
        # Generate predictions day by day
        for day in range(days_ahead):
            # Move to next business day (skip weekends)
            current_date = current_date + timedelta(days=1)
            while current_date.weekday() >= 5:  # Skip Saturday(5) and Sunday(6)
                current_date = current_date + timedelta(days=1)
            
            predictions['dates'].append(current_date)
            
            # Prepare features for prediction
            features = last_data[self.feature_columns].values
            features_scaled = self.feature_scaler.transform(features)
            
            # Get predictions from both models
            lr_pred = self.models['Linear_Regression'].predict(features_scaled)[0]
            rf_pred = self.models['Random_Forest'].predict(features_scaled)[0]
            
            predictions['Linear_Regression'].append(lr_pred)
            predictions['Random_Forest'].append(rf_pred)
            
            # Update features for next prediction
            # Shift lag features
            last_data['Close_lag_2'] = last_data['Close_lag_1'].values[0]
            last_data['Close_lag_1'] = lr_pred  # Use LR prediction as next input
            
            # Update moving averages (simplified)
            last_data['MA_5'] = lr_pred
            last_data['MA_10'] = lr_pred
            last_data['MA_20'] = lr_pred
            
            # Update price change
            last_data['Price_Change'] = (lr_pred - last_data['Close_lag_1'].values[0]) / last_data['Close_lag_1'].values[0]
            
            # Print progress every 20 days
            if (day + 1) % 20 == 0:
                print(f"    Predicted {day + 1}/{days_ahead} days...")
        
        # Create DataFrame with predictions
        future_df = pd.DataFrame({
            'Date': predictions['dates'],
            'Linear_Regression_Prediction': predictions['Linear_Regression'],
            'Random_Forest_Prediction': predictions['Random_Forest'],
            'Average_Prediction': [
                (lr + rf) / 2 
                for lr, rf in zip(predictions['Linear_Regression'], 
                                 predictions['Random_Forest'])
            ]
        })
        
        print(f"\n✅ Generated {days_ahead} future predictions!")
        print(f" Prediction range: {future_df['Date'].iloc[0].date()} to {future_df['Date'].iloc[-1].date()}")
        print(f" Predicted price range: ${future_df['Average_Prediction'].min():.2f} - ${future_df['Average_Prediction'].max():.2f}")
        
        return future_df
    
    def export_to_excel(self, future_predictions, filename='AAPL_Future_Predictions.xlsx'):
        """Step 4: Export predictions to professional Excel file"""
        print("\n" + "=" * 70)
        print("STEP 4: EXPORTING TO EXCEL")
        print("=" * 70)
        
        print(f"\n Creating Excel file: {filename}")
        
        # Create Excel writer
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        
        # Sheet 1: Summary & Key Metrics
        print(" Creating Summary sheet...")
        summary_data = {
            'Metric': [
                'Stock Symbol',
                'Analysis Date',
                'Historical Data Period',
                'Prediction Period',
                'Number of Predictions',
                'Starting Price',
                'Predicted Final Price (Average)',
                'Predicted Lowest Price',
                'Predicted Highest Price',
                'Linear Regression Accuracy',
                'Random Forest Accuracy'
            ],
            'Value': [
                self.symbol,
                datetime.now().strftime('%Y-%m-%d'),
                f"5 years ({len(self.data)} days)",
                f"{future_predictions['Date'].iloc[0].date()} to {future_predictions['Date'].iloc[-1].date()}",
                len(future_predictions),
                f"${self.data['Close'].iloc[-1]:.2f}",
                f"${future_predictions['Average_Prediction'].iloc[-1]:.2f}",
                f"${future_predictions['Average_Prediction'].min():.2f}",
                f"${future_predictions['Average_Prediction'].max():.2f}",
                'See Model Performance sheet',
                'See Model Performance sheet'
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Future Predictions (Main Results)
        print("📄 Creating Future Predictions sheet...")
        future_export = future_predictions.copy()
        future_export['Date'] = future_export['Date'].dt.strftime('%Y-%m-%d')
        
        # Round prices to 2 decimals
        for col in ['Linear_Regression_Prediction', 'Random_Forest_Prediction', 'Average_Prediction']:
            future_export[col] = future_export[col].round(2)
        
        future_export.to_excel(writer, sheet_name='Future Predictions', index=False)
        
        # Sheet 3: Weekly Summary
        print(" Creating Weekly Summary sheet...")
        future_predictions['Week'] = (future_predictions.index // 5) + 1
        weekly_summary = future_predictions.groupby('Week').agg({
            'Date': 'last',
            'Linear_Regression_Prediction': 'mean',
            'Random_Forest_Prediction': 'mean',
            'Average_Prediction': 'mean'
        }).reset_index()
        
        weekly_summary.columns = ['Week Number', 'End Date', 'LR Avg Price', 
                                  'RF Avg Price', 'Average Price']
        weekly_summary['End Date'] = weekly_summary['End Date'].dt.strftime('%Y-%m-%d')
        
        for col in ['LR Avg Price', 'RF Avg Price', 'Average Price']:
            weekly_summary[col] = weekly_summary[col].round(2)
        
        weekly_summary.to_excel(writer, sheet_name='Weekly Summary', index=False)
        
        # Sheet 4: Historical Data (Last 100 days)
        print(" Creating Historical Data sheet...")
        historical = self.data[['Date', 'Open', 'High', 'Low', 'Close', 'Volume']].tail(100).copy()
        historical['Date'] = historical['Date'].dt.strftime('%Y-%m-%d')
        historical.to_excel(writer, sheet_name='Historical Data', index=False)
        
        # Sheet 5: Model Performance
        print(" Creating Model Performance sheet...")
        # This would contain your model metrics - add after training
        performance_data = {
            'Model': ['Linear Regression', 'Random Forest'],
            'Type': ['Statistical Model', 'Machine Learning Ensemble'],
            'Advantages': [
                'Fast, interpretable, good for linear trends',
                'Handles non-linearity, robust to outliers'
            ],
            'Status': ['Trained & Ready', 'Trained & Ready']
        }
        df_performance = pd.DataFrame(performance_data)
        df_performance.to_excel(writer, sheet_name='Model Performance', index=False)
        
        writer.close()
        
        # Format the Excel file
        print("🎨 Applying professional formatting...")
        self._format_excel(filename, future_predictions)
        
        print(f"\n✅ Excel file created successfully!")
        print(f" Location: {filename}")
        print(f" Contains 5 sheets:")
        print(f"   1. Summary - Key metrics and overview")
        print(f"   2. Future Predictions - All {len(future_predictions)} daily predictions")
        print(f"   3. Weekly Summary - Aggregated weekly forecasts")
        print(f"   4. Historical Data - Last 100 days of actual prices")
        print(f"   5. Model Performance - Information about models used")
        
    def _format_excel(self, filename, future_predictions):
        """Apply professional formatting to Excel file"""
        wb = openpyxl.load_workbook(filename)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells with data
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                   min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
        
        # Add chart to Future Predictions sheet
        ws = wb['Future Predictions']
        chart = LineChart()
        chart.title = "AAPL Stock Price Forecast - Next 100 Days"
        chart.style = 13
        chart.y_axis.title = 'Price ($)'
        chart.x_axis.title = 'Trading Days'
        chart.height = 10
        chart.width = 20
        
        # Add data to chart
        data = Reference(ws, min_col=4, min_row=1, max_row=min(101, ws.max_row))
        chart.add_data(data, titles_from_data=True)
        
        ws.add_chart(chart, "F2")
        
        wb.save(filename)

def main():
    """
    Main execution: Complete prediction pipeline
    """
    print("\n" + "=" * 70)
    print(" AAPL STOCK PRICE PREDICTION - NEXT 100 DAYS")
    print("=" * 70)
    print(f" Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Initialize predictor
    predictor = FuturePricePredictor("AAPL")
    
    # Step 1: Prepare historical data
    predictor.prepare_data(period="5y")
    
    # Step 2: Train prediction models
    model_performance = predictor.train_models()
    
    # Step 3: Predict next 100 days
    future_predictions = predictor.predict_future(days_ahead=100)
    
    # Display sample predictions
    print("\n SAMPLE PREDICTIONS (First 10 days):")
    print("-" * 70)
    print(future_predictions.head(10).to_string(index=False))
    
    print("\n SAMPLE PREDICTIONS (Last 10 days):")
    print("-" * 70)
    print(future_predictions.tail(10).to_string(index=False))
    
    # Step 4: Export to Excel
    predictor.export_to_excel(future_predictions, 'AAPL_Future_Predictions.xlsx')
    
    # Final summary
    print("\n" + "=" * 70)
    print(" PREDICTION COMPLETE!")
    print("=" * 70)
    print(f"✅ Analyzed 5 years of historical data")
    print(f"✅ Trained 2 prediction models")
    print(f"✅ Generated 100 days of future predictions")
    print(f"✅ Exported results to Excel")
    print(f"\n📈 Key Insights:")
    print(f"   Starting Price: ${predictor.data['Close'].iloc[-1]:.2f}")
    print(f"   Predicted Price (Day 100): ${future_predictions['Average_Prediction'].iloc[-1]:.2f}")
    print(f"   Price Change: ${future_predictions['Average_Prediction'].iloc[-1] - predictor.data['Close'].iloc[-1]:.2f}")
    print(f"   Percentage Change: {((future_predictions['Average_Prediction'].iloc[-1] / predictor.data['Close'].iloc[-1]) - 1) * 100:.2f}%")
    
    # print("\n Next Steps:")
    # print("   1. Open AAPL_Future_Predictions.xlsx")
    # print("   2. Review the 'Summary' sheet for key metrics")
    # print("   3. Check 'Future Predictions' sheet for daily forecasts")
    # print("   4. Use data for your presentation/report")
    
    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()